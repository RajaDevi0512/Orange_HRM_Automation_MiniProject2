"""
This file contains read and write data from excel file

"""

from openpyxl import load_workbook

class ExcelFunction:
    def __init__(self,file_name, sheet_name):
        self.file = file_name
        self.sheet = sheet_name

    # Checks the row count
    def row_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_row
    
    # Checks the column count
    def column_count(self):
        workbook = load_workbook(self.file)
        sheet =workbook[self.sheet]
        return sheet.max_column
    
    # Reads the data for login
    def read_data(self, row_number, column_number):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.cell(row=row_number, column=column_number).value
    
    # Write the test status into the excel sheet
    def write_data(self,row_number, column_number, data):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        sheet.cell(row = row_number, column = column_number).value = data
        workbook.save(self.file)